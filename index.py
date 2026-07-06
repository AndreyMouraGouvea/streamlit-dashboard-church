import os
import random
import pandas as pd
from datetime import date, timedelta
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from win32com import client  


POSTOS = ["Porta", "Rua", "Banheiro (MASC)", "Púlpito"]

DIA_SEMANA_PT = {
    0: "Segunda",
    1: "Terça",
    2: "Quarta",
    3: "Quinta",
    4: "Sexta",
    5: "Sábado",
    6: "Domingo",
}


def nome_mes(m):
    meses = [
        "",
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    return meses[m].upper()


def dias_terca_sabado(ano: int, mes: int):
    d = date(ano, mes, 1)
    dias = []
    while d.month == mes:
        if d.weekday() in (1, 5):  # 1 = Terça, 5 = Sábado
            dias.append(d)
        d += timedelta(days=1)
    return dias


def carregar_indisponibilidade(caminho_entrada: str):
    df = pd.read_excel(caminho_entrada)

    if "Nome" not in df.columns or "Data Indisponível" not in df.columns:
        raise ValueError(
            "O arquivo deve ter as colunas 'Nome' e 'Data Indisponível'."
        )

    df = df[["Nome", "Data Indisponível"]].dropna(subset=["Nome"])

    df["Nome"] = df["Nome"].astype(str).str.strip()
    df["Data Indisponível"] = pd.to_datetime(
        df["Data Indisponível"], errors="coerce"
    ).dt.date

    nomes = sorted(
        {n for n in df["Nome"].unique() if isinstance(n, str) and n.strip()}
    )

    indisponiveis = {nome: set() for nome in nomes}
    for _, linha in df.iterrows():
        nome = linha["Nome"]
        dia = linha["Data Indisponível"]
        if pd.isna(dia):
            continue
        indisponiveis.setdefault(nome, set()).add(dia)

    return nomes, indisponiveis


def montar_linhas_escala(nomes, indisponiveis, ano, mes):
    dias_trabalho = dias_terca_sabado(ano, mes)
    if not dias_trabalho:
        raise ValueError("Não há terças ou sábados nesse mês/ano.")

    linhas_escala = []
    contagem = {nome: 0 for nome in nomes}

    for dia in dias_trabalho:
        disponiveis = [n for n in nomes if dia not in indisponiveis.get(n, set())]
        disponiveis.sort(key=lambda n: contagem[n])

        grupos = {}
        for n in disponiveis:
            grupos.setdefault(contagem[n], []).append(n)
        embaralhados = []
        for c in sorted(grupos):
            random.shuffle(grupos[c])
            embaralhados.extend(grupos[c])

        atribuicoes = embaralhados[: len(POSTOS)]

        for i, posto in enumerate(POSTOS):
            nome_escolhido = atribuicoes[i] if i < len(atribuicoes) else ""
            if nome_escolhido:
                contagem[nome_escolhido] += 1

            linhas_escala.append(
                {
                    "Data": dia,
                    "DiaSemana": DIA_SEMANA_PT[dia.weekday()],
                    "Posto": posto,
                    "Nome": nome_escolhido,
                }
            )

    return linhas_escala


def aplicar_borda_intervalo(ws, range_ref, border):
    for row in ws[range_ref]:
        for cell in row:
            cell.border = border


def criar_excel_template(linhas_escala, caminho_saida, ano, mes):
    linhas_por_data = []
    data_atual = None
    linha_corrente = None

    for item in linhas_escala:
        d = item["Data"]
        if data_atual != d:
            if linha_corrente is not None:
                linhas_por_data.append(linha_corrente)
            data_atual = d
            linha_corrente = {
                "Data": d,
                "DiaSemana": item["DiaSemana"],
                "Porta": "",
                "Rua": "",
                "Banheiro (MASC)": "",
                "Púlpito": "",
            }
        linha_corrente[item["Posto"]] = item["Nome"]

    if linha_corrente is not None:
        linhas_por_data.append(linha_corrente)

    wb = Workbook()
    ws = wb.active
    ws.title = "Escala"

    azul_escuro = "002D5C"
    azul_header = "003E7E"
    branco = "FFFFFF"
    borda_fina = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    # Título 1
    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value = "ESCALA DE TRABALHO"
    c.font = Font(size=18, bold=True, color=azul_escuro)
    c.alignment = Alignment(horizontal="center", vertical="center")
    aplicar_borda_intervalo(ws, "B1:F1", borda_fina)

    # Título 2
    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = f"{nome_mes(mes)}/{ano}"
    c.font = Font(size=14, bold=True, color=azul_escuro)
    c.alignment = Alignment(horizontal="center", vertical="center")
    aplicar_borda_intervalo(ws, "B2:F2", borda_fina)

    # Título 3
    ws.merge_cells("B3:F3")
    c = ws["B3"]
    c.value = "TERÇAS E SÁBADOS"
    c.font = Font(size=12, bold=True, color=azul_escuro)
    c.alignment = Alignment(horizontal="center", vertical="center")
    aplicar_borda_intervalo(ws, "B3:F3", borda_fina)

    header_row = 4
    headers = ["DATA", "PORTA", "RUA", "BANHEIRO (MASC)", "PÚLPITO"]
    for col_idx, texto in enumerate(headers, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=texto)
        cell.font = Font(bold=True, color=branco)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=azul_header)
        cell.border = borda_fina

    row = header_row + 1
    for item in linhas_por_data:
        data = item["Data"]
        dia_semana = item["DiaSemana"].upper()
        texto_data = data.strftime("%d/%m/%Y") + "\n" + dia_semana

        cell_data = ws.cell(row=row, column=2, value=texto_data)
        cell_data.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell_data.border = borda_fina

        cell_porta = ws.cell(row=row, column=3, value=item.get("Porta", ""))
        cell_porta.alignment = Alignment(horizontal="center", vertical="center")
        cell_porta.border = borda_fina

        cell_rua = ws.cell(row=row, column=4, value=item.get("Rua", ""))
        cell_rua.alignment = Alignment(horizontal="center", vertical="center")
        cell_rua.border = borda_fina

        cell_banheiro = ws.cell(
            row=row, column=5, value=item.get("Banheiro (MASC)", "")
        )
        cell_banheiro.alignment = Alignment(horizontal="center", vertical="center")
        cell_banheiro.border = borda_fina

        cell_pulpito = ws.cell(row=row, column=6, value=item.get("Púlpito", ""))
        cell_pulpito.alignment = Alignment(horizontal="center", vertical="center")
        cell_pulpito.border = borda_fina

        row += 1

    wb.save(caminho_saida)


def gerar_pdf(caminho_excel: str, caminho_pdf: str):
    excel = client.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(caminho_excel))
        ws = wb.Worksheets(1)

        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesTall = 1
        ws.PageSetup.FitToPagesWide = 1

        ws.ExportAsFixedFormat(0, os.path.abspath(caminho_pdf))
    finally:
        wb.Close(SaveChanges=False)
        excel.Quit()


def gerar_escala(caminho_entrada: str, caminho_saida: str, ano: int, mes: int):
    nomes, indisponiveis = carregar_indisponibilidade(caminho_entrada)
    if not nomes:
        raise ValueError("Nenhum nome encontrado no arquivo de entrada.")

    linhas_escala = montar_linhas_escala(nomes, indisponiveis, ano, mes)
    criar_excel_template(linhas_escala, caminho_saida, ano, mes)

    pasta = os.path.dirname(caminho_saida) or "."
    nome_base, _ = os.path.splitext(os.path.basename(caminho_saida))
    caminho_pdf = os.path.join(pasta, f"{nome_base}.pdf")

    gerar_pdf(caminho_saida, caminho_pdf)


def selecionar_planilha(entry_caminho):
    caminho = filedialog.askopenfilename(
        title="Selecione a planilha de entrada",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
    )
    if caminho:
        entry_caminho.delete(0, "end")
        entry_caminho.insert(0, caminho)


def acao_gerar_escala(entry_caminho, entry_ano, entry_mes):
    try:
        caminho_entrada = entry_caminho.get().strip()
        ano_texto = entry_ano.get().strip()
        mes_texto = entry_mes.get().strip()

        if not caminho_entrada:
            messagebox.showerror("Erro", "Selecione a planilha de entrada.")
            return

        if not ano_texto.isdigit() or not mes_texto.isdigit():
            messagebox.showerror("Erro", "Ano e mês devem ser numéricos.")
            return

        ano = int(ano_texto)
        mes = int(mes_texto)
        if mes < 1 or mes > 12:
            messagebox.showerror("Erro", "O mês deve estar entre 1 e 12.")
            return

        pasta = os.path.dirname(caminho_entrada) or "."
        caminho_saida = os.path.join(
            pasta, f"ESCALA_CCB_{nome_mes(mes)}_{ano}.xlsx"
        )

        gerar_escala(caminho_entrada, caminho_saida, ano, mes)
        messagebox.showinfo(
            "Sucesso",
            f"Escala gerada com sucesso em:\n{caminho_saida}\n\n"
        )
    except Exception as e:
        messagebox.showerror("Erro ao gerar escala", str(e))


def main():
    ctk.set_appearance_mode("system")
    ctk.set_default_color_theme("blue")

    app = ctk.CTk()
    app.title("Gerador de Escala")
    app.geometry("500x230")

    frame_entrada = ctk.CTkFrame(app)
    frame_entrada.pack(padx=10, pady=10, fill="x")

    lbl_entrada = ctk.CTkLabel(frame_entrada, text="Planilha de entrada:")
    lbl_entrada.pack(anchor="w", padx=5, pady=2)

    subframe_entrada = ctk.CTkFrame(frame_entrada)
    subframe_entrada.pack(fill="x", padx=5, pady=2)

    entry_caminho = ctk.CTkEntry(subframe_entrada)
    entry_caminho.pack(side="left", fill="x", expand=True, padx=(0, 5))

    btn_selecionar = ctk.CTkButton(
        subframe_entrada,
        text="Selecionar...",
        command=lambda: selecionar_planilha(entry_caminho)
    )
    btn_selecionar.pack(side="left")

    frame_data = ctk.CTkFrame(app)
    frame_data.pack(padx=10, pady=5, fill="x")

    lbl_ano = ctk.CTkLabel(frame_data, text="Ano:")
    lbl_ano.grid(row=0, column=0, padx=5, pady=5, sticky="w")
    entry_ano = ctk.CTkEntry(frame_data, width=80)
    entry_ano.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    lbl_mes = ctk.CTkLabel(frame_data, text="Mês (1-12):")
    lbl_mes.grid(row=0, column=2, padx=5, pady=5, sticky="w")
    entry_mes = ctk.CTkEntry(frame_data, width=80)
    entry_mes.grid(row=0, column=3, padx=5, pady=5, sticky="w")

    btn_gerar = ctk.CTkButton(
        app,
        text="Gerar escala",
        command=lambda: acao_gerar_escala(
            entry_caminho, entry_ano, entry_mes
        )
    )
    btn_gerar.pack(pady=15)

    app.mainloop()


if __name__ == "__main__":
    main()