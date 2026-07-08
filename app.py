import random
import streamlit as st
import pandas as pd
from datetime import date, timedelta, datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fpdf import FPDF
from io import BytesIO
import os

st.set_page_config(page_title="Gerador de Escala", layout="wide")
st.set_option("client.showErrorDetails", True)

DIA_SEMANA_PT = {
    0: "Segunda", 1: "Terça", 2: "Quarta", 3: "Quinta",
    4: "Sexta", 5: "Sábado", 6: "Domingo",
}

POSTOS_PADRAO = ["Porta", "Rua", "Banheiro (MASC)", "Púlpito"]


def nome_mes(m: int) -> str:
    meses = [
        "", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
    ]
    return meses[m].upper()


def gerar_dias(ano: int, mes: int, weekdays: list[int]) -> list[date]:
    d = date(ano, mes, 1)
    dias = []
    while d.month == mes:
        if d.weekday() in weekdays:
            dias.append(d)
        d += timedelta(days=1)
    return dias


def parse_datas(texto: str) -> set[date]:
    if not texto or not texto.strip():
        return set()
    datas = set()
    for linha in texto.strip().split("\n"):
        linha = linha.strip()
        if not linha:
            continue
        try:
            datas.add(datetime.strptime(linha, "%d/%m/%Y").date())
        except ValueError:
            st.error(f"Data inválida: '{linha}' — use dd/mm/aaaa")
    return datas


def montar_escala(
    nomes: list[str],
    indisponiveis: dict[str, set[date]],
    dias: list[date],
    postos: list[str],
) -> list[dict]:
    linhas = []
    contagem = {nome: 0 for nome in nomes}

    for dia in dias:
        disponiveis = [p for p in nomes if dia not in indisponiveis.get(p, set())]
        disponiveis.sort(key=lambda n: contagem[n])

        grupos = {}
        for n in disponiveis:
            grupos.setdefault(contagem[n], []).append(n)
        embaralhados = []
        for c in sorted(grupos):
            random.shuffle(grupos[c])
            embaralhados.extend(grupos[c])

        atribuicoes = embaralhados[: len(postos)]

        for i, posto in enumerate(postos):
            nome = atribuicoes[i] if i < len(atribuicoes) else ""
            if nome:
                contagem[nome] += 1
            linhas.append({
                "Data": dia,
                "DiaSemana": DIA_SEMANA_PT[dia.weekday()],
                "Posto": posto,
                "Nome": nome,
            })

    return linhas


def borda_fina() -> Border:
    return Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )


def agrupar_por_data(linhas_escala: list[dict], postos: list[str]) -> list[dict]:
    grupos = []
    corrente = None
    data_atual = None
    for item in linhas_escala:
        d = item["Data"]
        if data_atual != d:
            if corrente:
                grupos.append(corrente)
            data_atual = d
            corrente = {"Data": d, "DiaSemana": item["DiaSemana"]}
            for p in postos:
                corrente[p] = ""
        corrente[item["Posto"]] = item["Nome"]
    if corrente:
        grupos.append(corrente)
    return grupos


def criar_excel_bytes(
    linhas_escala: list[dict],
    postos: list[str],
    ano: int,
    mes: int,
    weekdays: list[int],
) -> bytes:
    grupos = agrupar_por_data(linhas_escala, postos)

    wb = Workbook()
    ws = wb.active
    ws.title = "Escala"

    AZUL = "002D5C"
    AZUL_H = "003E7E"
    BRANCO = "FFFFFF"
    BORDA = borda_fina()

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    for idx in range(len(postos)):
        ws.column_dimensions[get_column_letter(3 + idx)].width = 22

    last_col = 2 + len(postos)

    for r in range(1, 4):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
        for c in range(2, last_col + 1):
            ws.cell(row=r, column=c).border = BORDA

    c = ws.cell(row=1, column=2, value="ESCALA DE TRABALHO")
    c.font = Font(size=18, bold=True, color=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")

    c = ws.cell(row=2, column=2, value=f"{nome_mes(mes)}/{ano}")
    c.font = Font(size=14, bold=True, color=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")

    dias_texto = " \u00b7 ".join(DIA_SEMANA_PT[d] for d in sorted(weekdays))
    c = ws.cell(row=3, column=2, value=dias_texto)
    c.font = Font(size=12, bold=True, color=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")

    header_row = 4
    headers = ["DATA"] + [p.upper() for p in postos]
    for ci, texto in enumerate(headers, start=2):
        cell = ws.cell(row=header_row, column=ci, value=texto)
        cell.font = Font(bold=True, color=BRANCO)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=AZUL_H)
        cell.border = BORDA

    row = header_row + 1
    for item in grupos:
        txt = item["Data"].strftime("%d/%m/%Y") + "\n" + item["DiaSemana"].upper()
        cell = ws.cell(row=row, column=2, value=txt)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDA

        for ci, posto in enumerate(postos, start=3):
            cell = ws.cell(row=row, column=ci, value=item.get(posto, ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDA

        row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def criar_pdf_bytes(
    linhas_escala: list[dict],
    postos: list[str],
    ano: int,
    mes: int,
    weekdays: list[int],
) -> bytes:
    grupos = agrupar_por_data(linhas_escala, postos)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)

    W = 297
    col_w = [48] + [35] * len(postos)
    total_w = sum(col_w)
    x0 = (W - total_w) / 2

    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "ESCALA DE TRABALHO", ln=1, align="C")
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, f"{nome_mes(mes)}/{ano}", ln=1, align="C")
    dias_texto = " / ".join(DIA_SEMANA_PT[d] for d in sorted(weekdays))
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 8, dias_texto, ln=1, align="C")
    pdf.ln(5)

    def cabecalho():
        pdf.set_x(x0)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(0, 62, 126)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(["DATA"] + [p.upper() for p in postos]):
            pdf.cell(col_w[i], 8, h, border=1, align="C", fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

    cabecalho()

    for item in grupos:
        if pdf.get_y() + 8 > 190:
            pdf.add_page()
            cabecalho()

        txt = f"{item['Data'].strftime('%d/%m/%Y')} ({item['DiaSemana'].upper()})"
        pdf.set_x(x0)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(col_w[0], 7, txt, border=1, align="C")
        for i, posto in enumerate(postos):
            pdf.cell(col_w[i + 1], 7, item.get(posto, ""), border=1, align="C")
        pdf.ln()

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf.read()


# ── Session state ──
for key, default in [("num_postos", 4), ("num_pessoas", 4)]:
    if key not in st.session_state:
        st.session_state[key] = default


def remover_posto(idx: int):
    valores = [st.session_state.get(f"posto_{j}", "") for j in range(st.session_state.num_postos)]
    valores.pop(idx)
    for j in range(st.session_state.num_postos):
        st.session_state.pop(f"posto_{j}", None)
    for j, val in enumerate(valores):
        st.session_state[f"posto_{j}"] = val
    st.session_state.num_postos -= 1


def remover_pessoa(idx: int):
    nomes = [st.session_state.get(f"pessoa_nome_{j}", "") for j in range(st.session_state.num_pessoas)]
    datas = [st.session_state.get(f"pessoa_datas_{j}", "") for j in range(st.session_state.num_pessoas)]
    nomes.pop(idx)
    datas.pop(idx)
    for j in range(st.session_state.num_pessoas):
        st.session_state.pop(f"pessoa_nome_{j}", None)
        st.session_state.pop(f"pessoa_datas_{j}", None)
    for j, (nome, data) in enumerate(zip(nomes, datas)):
        st.session_state[f"pessoa_nome_{j}"] = nome
        st.session_state[f"pessoa_datas_{j}"] = data
    st.session_state.num_pessoas -= 1


# ── UI ──
st.title("Gerador de Escala")

with st.sidebar:
    st.header("Configuração")
    ano = st.number_input("Ano", min_value=2000, max_value=2100, value=2026)
    mes = st.number_input("M\u00eas", min_value=1, max_value=12, value=6)

    st.subheader("Dias da semana")
    weekdays = [i for i in range(7) if st.checkbox(DIA_SEMANA_PT[i], value=(i in (1, 5)))]

st.header("Postos (cargos/funções)")
postos = []
for i in range(st.session_state.num_postos):
    c1, c2 = st.columns([5, 1])
    with c1:
        val = st.text_input(
            f"Posto {i + 1}",
            value=POSTOS_PADRAO[i] if i < len(POSTOS_PADRAO) else "",
            key=f"posto_{i}",
        )
        postos.append(val.strip())
    with c2:
        st.button(
            "✕",
            key=f"rm_posto_{i}",
            disabled=st.session_state.num_postos <= 1,
            on_click=remover_posto,
            args=(i,),
        )

if st.button("+ Adicionar posto", key="btn_add_posto", use_container_width=True):
    st.session_state.num_postos += 1
    st.rerun()

st.divider()
st.header("Pessoas")

for i in range(st.session_state.num_pessoas):
    with st.container(border=True):
        c1, c2 = st.columns([5, 1])
        with c1:
            st.text_input("Nome", key=f"pessoa_nome_{i}")
        with c2:
            st.button(
                "✕",
                key=f"rm_pessoa_{i}",
                disabled=st.session_state.num_pessoas <= 1,
                on_click=remover_pessoa,
                args=(i,),
            )
        st.text_area(
            "Datas indisponíveis (dd/mm/aaaa, uma por linha)",
            height=70,
            key=f"pessoa_datas_{i}",
        )

if st.button("+ Adicionar pessoa", key="btn_add_pessoa", use_container_width=True):
    st.session_state.num_pessoas += 1
    st.rerun()

st.divider()

# ── Generate ──
if st.button("Gerar Escala", type="primary", use_container_width=True):
    postos_val = [p for p in postos if p]
    if not postos_val:
        st.error("Adicione pelo menos um posto.")
        st.stop()

    if not weekdays:
        st.error("Selecione pelo menos um dia da semana.")
        st.stop()

    nomes = []
    indisp = {}
    for i in range(st.session_state.num_pessoas):
        nome = st.session_state.get(f"pessoa_nome_{i}", "").strip()
        if not nome:
            continue
        nomes.append(nome)
        indisp[nome] = parse_datas(st.session_state.get(f"pessoa_datas_{i}", ""))

    if not nomes:
        st.error("Adicione pelo menos uma pessoa.")
        st.stop()

    if len(nomes) < len(postos_val):
        st.error(
            f"Você precisa de pelo menos {len(postos_val)} pessoas "
            f"para cobrir todos os postos (você tem {len(nomes)})."
        )
        st.stop()

    dias = gerar_dias(ano, mes, weekdays)
    if not dias:
        st.error(
            f"Não há {', '.join(DIA_SEMANA_PT[d] for d in weekdays)} "
            f"em {nome_mes(mes)}/{ano}."
        )
        st.stop()

    escala = montar_escala(nomes, indisp, dias, postos_val)

    st.success(
        f"Escala gerada! {len(dias)} dia(s), "
        f"{len(postos_val)} posto(s), {len(nomes)} pessoa(s)."
    )

    df = pd.DataFrame(escala)
    pivot = df.pivot_table(
        index=["Data", "DiaSemana"],
        columns="Posto",
        values="Nome",
        aggfunc="first",
    ).reset_index().fillna("")
    pivot.columns.name = None
    pivot["Data"] = pivot["Data"].apply(lambda d: d.strftime("%d/%m/%Y"))
    st.dataframe(pivot, use_container_width=True, hide_index=True)

    xlsx = criar_excel_bytes(escala, postos_val, ano, mes, weekdays)
    pdf = criar_pdf_bytes(escala, postos_val, ano, mes, weekdays)

    base = f"ESCALA_{nome_mes(mes)}_{ano}"

    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "Download Excel (.xlsx)",
            data=xlsx,
            file_name=f"{base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with c2:
        st.download_button(
            "Download PDF (.pdf)",
            data=pdf,
            file_name=f"{base}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
